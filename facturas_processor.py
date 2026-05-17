import urllib.request
import urllib.error
import json
import os
import urllib.parse
import sqlite3
import hashlib
import logging
import time
import shutil
import xml.etree.ElementTree as ET
from datetime import datetime
from logging.handlers import RotatingFileHandler
import re

# ══════════════════════════════════════════════════════════════
# PATCH 1 — PyPDF2 → pypdf
# PyPDF2 está deprecado desde 2022. pypdf es su sucesor oficial,
# con mejor extracción de texto y soporte activo.
# Instalar: pip install pypdf
# ══════════════════════════════════════════════════════════════
try:
    from pypdf import PdfReader          # pypdf (nuevo)
    _PDF_BACKEND = "pypdf"
except ImportError:
    import PyPDF2                        # fallback si aún no se migró
    _PDF_BACKEND = "PyPDF2"

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────
# CONFIGURACION
# ─────────────────────────────────────────
API_KEY          = os.environ.get("OPENROUTER_API_KEY")
TELEGRAM_TOKEN   = os.environ.get("TELEGRAM_TOKEN")
TELEGRAM_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID")

CARPETA_CLIENTES = os.path.expanduser("~/.private_data/clientes")

MODELOS = [
    "meta-llama/llama-3.3-70b-instruct:free",
    "nvidia/nemotron-3-super-120b-a12b:free",
    "nvidia/nemotron-3-nano-30b-a3b:free",
    "nvidia/nemotron-nano-12b-2-VL:free",
    "openai/gpt-oss-120b:free",
    "openai/gpt-oss-20b:free",
    "google/gemma-4-26b-a4b-it:free",
    "nousresearch/hermes-3-llama-3.1-405b:free",
    "openrouter/free",
]

CAMPOS_OBLIGATORIOS = ["proveedor", "monto_total", "moneda"]

# ══════════════════════════════════════════════════════════════
# PATCH 2 — LOG INJECTION SANITIZACIÓN
# Un atacante puede subir un archivo llamado:
#   "factura\n[CRITICAL] admin logged in.pdf"
# e inyectar entradas falsas en los logs.
# Solución: sanitizar cualquier string antes de loggearlo.
# ══════════════════════════════════════════════════════════════
def _sanitize_log(value: str) -> str:
    """
    Elimina saltos de línea, tabs y caracteres de control del string
    para prevenir log injection. Trunca a 500 chars para evitar logs enormes.
    """
    if not isinstance(value, str):
        value = str(value)
    # Reemplazar cualquier carácter de control (incluye \n, \r, \t)
    sanitized = re.sub(r'[\x00-\x1f\x7f]', ' ', value)
    return sanitized[:500]

def slog(level: str, msg: str, *args):
    """
    Wrapper de logging con sanitización automática.
    Uso igual que logging.info/warning/error:
      slog("info",  "Factura procesada: {}", archivo)
      slog("error", "Fallo en {}/{}", cliente, archivo)
    """
    # Sanitizar todos los argumentos interpolados
    safe_args = [_sanitize_log(a) for a in args]
    safe_msg  = _sanitize_log(msg).format(*safe_args) if safe_args else _sanitize_log(msg)
    getattr(logging, level)(safe_msg)

# ── Logging con rotación ─────────────────────────────────────
_handler = RotatingFileHandler(
    "procesamiento.log", maxBytes=2 * 1024 * 1024, backupCount=3, encoding="utf-8"
)
_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
logging.getLogger().addHandler(_handler)
logging.getLogger().setLevel(logging.INFO)

# ─────────────────────────────────────────
def enviar_telegram(mensaje):
    if not TELEGRAM_TOKEN or not TELEGRAM_CHAT_ID:
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    params = urllib.parse.urlencode({
        "chat_id": TELEGRAM_CHAT_ID,
        "text": mensaje,
        "parse_mode": "HTML"
    })
    for intento in range(3):
        try:
            req = urllib.request.Request(url + "?" + params)
            urllib.request.urlopen(req, timeout=15)
            return  # éxito, salir
        except Exception as e:
            slog("warning", "Telegram intento {}/3: {}", str(intento + 1), str(e))
            if intento < 2:
                time.sleep(10)
    slog("error", "Telegram falló después de 3 intentos")

# ─────────────────────────────────────────
def get_rutas_cliente(nombre_cliente):
    nombre_seguro = re.sub(r'[^\w\s\-.]', '', nombre_cliente).strip()
    if not nombre_seguro:
        raise ValueError("Nombre de cliente inválido")

    base_clientes = os.path.expanduser("~/.private_data/clientes")
    base = os.path.join(base_clientes, nombre_seguro)

    if not os.path.abspath(base).startswith(os.path.abspath(base_clientes)):
        raise ValueError("Ruta inválida - intento de path traversal")

    return {
        "base":       base,
        "facturas":   os.path.join(base, "facturas"),
        "procesados": os.path.join(base, "facturas", "procesados"),
        "error":      os.path.join(base, "facturas", "error"),
        "db":         os.path.join(base, "facturas.db"),
        "excel":      os.path.join(base, "resultados.xlsx"),
    }

def crear_carpetas_cliente(rutas):
    os.makedirs(rutas["facturas"],   exist_ok=True)
    os.makedirs(rutas["procesados"], exist_ok=True)
    os.makedirs(rutas["error"],      exist_ok=True)

def listar_clientes():
    if not os.path.exists(CARPETA_CLIENTES):
        os.makedirs(CARPETA_CLIENTES)
        return []
    return [
        f for f in os.listdir(CARPETA_CLIENTES)
        if os.path.isdir(os.path.join(CARPETA_CLIENTES, f))
    ]

# ─────────────────────────────────────────
CATEGORIAS_VALIDAS = [
    "Servicios", "Materiales y Suministros", "Transporte y Logistica",
    "Tecnologia y Software", "Nomina y RRHH", "Alquiler e Inmuebles",
    "Publicidad y Marketing", "Impuestos y Tasas", "Alimentacion", "Otros",
]

# ══════════════════════════════════════════════════════════════
# PATCH 3 — VALIDACIÓN DE CAMPOS FISCALES
# El validador original solo chequeaba que los campos existieran.
# Este agrega:
#   - Formato de RUC panameño (natural, jurídico, extranjero)
#   - Formato de fecha (DD/MM/AAAA o AAAA-MM-DD)
#   - Monto positivo y razonable (> 0, < 10 millones)
#   - Moneda válida
#   - CUFE: alfanumérico si existe
#   - Confianza: integer 0-100
# Retorna (bool, lista_de_advertencias) para loggear qué falló.
# ══════════════════════════════════════════════════════════════

# Patrones RUC Panamá:
# Natural:   8-XXX-XXXXX  (ej: 8-123-45678)
# Jurídico:  XXX-XXX-XXXXXX (ej: 155-12-34567)
# Extranjero: E-X-XXXXXX
# NITE:      N-XX-XXXXXX
# Simplificado: solo dígitos con guiones
_RUC_PATTERN = re.compile(
    r'^('
    r'\d{1,2}-\d{3,4}-\d{4,6}'      # Natural: 8-123-45678
    r'|\d{2,3}-\d{2,3}-\d{4,7}'     # Jurídico: 155-12-34567
    r'|[ENP]-\d{1,2}-\d{4,7}'       # Extranjero/NITE
    r'|\d{6,15}'                      # Solo dígitos (algunos sistemas)
    r')$'
)
_FECHA_PATTERNS = [
    re.compile(r'^\d{2}/\d{2}/\d{4}$'),   # DD/MM/AAAA
    re.compile(r'^\d{4}-\d{2}-\d{2}$'),   # AAAA-MM-DD (ISO, e-Tax XML)
]
_MONEDAS_VALIDAS = {"USD", "PAB", "EUR", "COP", "MXN", "PEN", "CLP", "ARS", "BRL"}

# ══════════════════════════════════════════════════════════════
# VALIDACIÓN DE RUC CONTRA DGI — portal e-Tax público
#
# La DGI Panamá expone una búsqueda pública en etax2.mef.gob.pa.
# Hacemos una request HTTP simple y parseamos la respuesta.
# Resultado se cachea en memoria (TTL 1 hora) para no martillar
# el portal en cada factura que llega.
# ══════════════════════════════════════════════════════════════
import urllib.parse as _uparse
import html as _html_lib

_dgi_cache: dict = {}          # {ruc: {"ok": bool, "nombre": str, "ts": float}}
_DGI_CACHE_TTL = 3600          # 1 hora

def validar_ruc_dgi(ruc: str) -> dict:
    """
    Consulta el portal e-Tax de la DGI para verificar si un RUC existe.

    Retorna dict con claves:
        ok      (bool)  — True si el RUC fue encontrado
        nombre  (str)   — Nombre del contribuyente según DGI (o "")
        error   (str)   — Mensaje de error si la consulta falló
        fuente  (str)   — "dgi" | "cache" | "error"
    """
    if not ruc:
        return {"ok": False, "nombre": "", "error": "RUC vacío", "fuente": "error"}

    ruc_limpio = re.sub(r'\s+', '', str(ruc)).upper()

    # Cache hit
    cached = _dgi_cache.get(ruc_limpio)
    if cached and (time.time() - cached["ts"]) < _DGI_CACHE_TTL:
        return {**cached, "fuente": "cache"}

    try:
        # Endpoint público del portal DGI / e-Tax 2.0
        url = "https://etax2.mef.gob.pa/etax2web/BusquedaContribuyente.xhtml"
        params = _uparse.urlencode({"ruc": ruc_limpio})
        full_url = f"{url}?{params}"

        req = urllib.request.Request(
            full_url,
            headers={
                "User-Agent": "Mozilla/5.0 (compatible; AutomatIA/1.0)",
                "Accept": "text/html,application/xhtml+xml",
            }
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            body = resp.read().decode("utf-8", errors="ignore")

        # Buscar nombre del contribuyente en la respuesta HTML
        # El portal muestra el nombre en un elemento con clase o texto reconocible
        nombre = ""

        # Patrón 1: tabla con datos del contribuyente
        m = re.search(
            r'(?:Nombre|Raz[oó]n Social|Contribuyente)[^<]*</[^>]+>\s*<[^>]+>\s*([^<]{3,120})',
            body, re.IGNORECASE
        )
        if m:
            nombre = _html_lib.unescape(m.group(1)).strip()

        # Patrón 2: valor en campo de texto
        if not nombre:
            m = re.search(
                r'<input[^>]+(?:nombre|razon)[^>]+value=["\']([^"\']{3,120})["\']',
                body, re.IGNORECASE
            )
            if m:
                nombre = _html_lib.unescape(m.group(1)).strip()

        # Patrón 3: texto libre después del RUC en la respuesta
        if not nombre:
            m = re.search(
                r'(?:' + re.escape(ruc_limpio) + r')[^<]{0,30}<[^>]+>\s*([A-ZÁÉÍÓÚÑ][^<]{2,80})',
                body, re.IGNORECASE
            )
            if m:
                nombre = _html_lib.unescape(m.group(1)).strip()

        # Determinar si el RUC existe: si encontramos nombre O si el HTML
        # no contiene mensajes de "no encontrado"
        no_encontrado_patterns = [
            "no se encontr", "not found", "ruc no existe",
            "no existe", "sin resultados", "no hay resultados"
        ]
        body_lower = body.lower()
        no_encontrado = any(p in body_lower for p in no_encontrado_patterns)

        # El RUC es válido si la página responde sin mensaje de error Y tiene contenido
        ruc_ok = bool(nombre) or (not no_encontrado and len(body) > 500)

        resultado = {
            "ok": ruc_ok,
            "nombre": nombre,
            "error": "" if ruc_ok else "RUC no encontrado en DGI",
            "ts": time.time(),
        }
        _dgi_cache[ruc_limpio] = resultado
        return {**resultado, "fuente": "dgi"}

    except urllib.error.URLError as e:
        slog("warning", "DGI no accesible para RUC {}: {}", ruc_limpio, str(e))
        return {"ok": None, "nombre": "", "error": f"DGI inaccesible: {e.reason}", "fuente": "error"}
    except Exception as e:
        slog("warning", "Error validando RUC {} en DGI: {}", ruc_limpio, str(e))
        return {"ok": None, "nombre": "", "error": str(e), "fuente": "error"}


# ══════════════════════════════════════════════════════════════
# CÁLCULO DE ITBMS ESPERADO
#
# Panamá aplica ITBMS del 7% sobre la base imponible (subtotal).
# Tasas especiales: 10% bebidas alcohólicas, 15% tabaco, 0% exentos.
# Esta función calcula el ITBMS esperado y detecta discrepancias.
# ══════════════════════════════════════════════════════════════
TASA_ITBMS_GENERAL   = 0.07
TASA_ITBMS_ALCOHOL   = 0.10
TASA_ITBMS_TABACO    = 0.15
TOLERANCIA_ITBMS     = 0.02   # ±2 centavos por redondeo

_KEYWORDS_ALCOHOL = re.compile(r'\b(cerveza|licor|ron|whisky|vino|alcohol|beer|spirits)\b', re.IGNORECASE)
_KEYWORDS_TABACO  = re.compile(r'\b(cigarro|cigarrillo|tabaco|tobacco|cigars?)\b', re.IGNORECASE)
_KEYWORDS_EXENTO  = re.compile(r'\b(exento|exent[ao]s?|exempt|medicamento|medicine|alimento básico)\b', re.IGNORECASE)

def calcular_itbms_esperado(subtotal: float, descripcion: str = "") -> tuple:
    """
    Calcula el ITBMS esperado dado un subtotal y descripción.
    Retorna (itbms_esperado: float, tasa: float, categoria: str)
    """
    desc = descripcion or ""
    if _KEYWORDS_EXENTO.search(desc):
        return 0.0, 0.0, "exento"
    elif _KEYWORDS_TABACO.search(desc):
        tasa = TASA_ITBMS_TABACO
        cat  = "tabaco_15pct"
    elif _KEYWORDS_ALCOHOL.search(desc):
        tasa = TASA_ITBMS_ALCOHOL
        cat  = "alcohol_10pct"
    else:
        tasa = TASA_ITBMS_GENERAL
        cat  = "general_7pct"

    return round(subtotal * tasa, 2), tasa, cat


def verificar_itbms(datos: dict) -> list:
    """
    Verifica que el ITBMS declarado sea consistente con el subtotal.
    Retorna lista de advertencias (vacía si todo está bien).
    """
    advertencias = []
    try:
        subtotal    = float(datos.get("subtotal") or 0)
        itbms       = datos.get("itbms")
        monto_total = float(datos.get("monto_total") or 0)
        descripcion = str(datos.get("descripcion") or "")

        if itbms is None:
            return advertencias  # no reportado, sin alarma

        itbms = float(itbms)

        if subtotal <= 0 and monto_total > 0:
            # Intentar reconstruir subtotal desde monto_total e itbms
            subtotal = monto_total - itbms

        if subtotal <= 0:
            return advertencias

        esperado, tasa, cat = calcular_itbms_esperado(subtotal, descripcion)

        diferencia = abs(itbms - esperado)
        if cat == "exento" and itbms > 0:
            advertencias.append(
                f"⚠ ITBMS: factura parece exenta pero declara ITBMS de B/.{itbms:.2f}"
            )
        elif diferencia > TOLERANCIA_ITBMS:
            pct_tasa = int(tasa * 100)
            advertencias.append(
                f"⚠ ITBMS posiblemente incorrecto: declarado B/.{itbms:.2f}, "
                f"esperado B/.{esperado:.2f} ({pct_tasa}% de B/.{subtotal:.2f}). "
                f"Diferencia: B/.{diferencia:.2f}"
            )

        # Verificar cuadratura: subtotal + itbms ≈ monto_total
        if monto_total > 0:
            suma = subtotal + itbms
            dif_total = abs(suma - monto_total)
            if dif_total > 0.05:
                advertencias.append(
                    f"⚠ Cuadratura: subtotal ({subtotal:.2f}) + ITBMS ({itbms:.2f}) "
                    f"= {suma:.2f} ≠ monto_total ({monto_total:.2f}). "
                    f"Diferencia: B/.{dif_total:.2f}"
                )
    except (TypeError, ValueError):
        pass

    return advertencias


def validar_campos_fiscales(datos: dict) -> tuple:
    """
    Valida campos fiscales panameños.
    Retorna (es_valido: bool, advertencias: list[str])
    """
    advertencias = []

    # ── Campos obligatorios ───────────────────────────────────
    for campo in CAMPOS_OBLIGATORIOS:
        if not datos.get(campo):
            advertencias.append(f"Campo obligatorio faltante: {campo}")

    if advertencias:
        return False, advertencias

    # ── monto_total: número positivo y razonable ──────────────
    try:
        monto = float(datos["monto_total"])
        if monto <= 0:
            advertencias.append(f"monto_total debe ser positivo, recibido: {monto}")
        elif monto > 10_000_000:
            # No bloquea, pero advierte — podría ser legítimo
            advertencias.append(f"monto_total inusualmente alto: {monto} (revisar manualmente)")
    except (ValueError, TypeError):
        advertencias.append(f"monto_total no es número: {datos.get('monto_total')}")
        return False, advertencias

    # ── moneda ────────────────────────────────────────────────
    moneda = str(datos.get("moneda", "")).strip().upper()
    if moneda not in _MONEDAS_VALIDAS:
        advertencias.append(f"Moneda no reconocida: '{moneda}' (se acepta, pero verificar)")

    # ── RUC (opcional pero si existe debe tener formato válido) ─
    ruc = datos.get("ruc")
    if ruc and ruc not in (None, "null", ""):
        ruc_limpio = str(ruc).strip()
        if not _RUC_PATTERN.match(ruc_limpio):
            advertencias.append(f"RUC con formato inusual: '{ruc_limpio}' (verificar)")

    # ── fecha (opcional pero si existe debe tener formato válido) ─
    fecha = datos.get("fecha")
    if fecha and fecha not in (None, "null", ""):
        fecha_str = str(fecha).strip()
        if not any(p.match(fecha_str) for p in _FECHA_PATTERNS):
            advertencias.append(f"Fecha con formato no estándar: '{fecha_str}' (esperado DD/MM/AAAA o AAAA-MM-DD)")

    # ── confianza: 0-100 ──────────────────────────────────────
    try:
        confianza = int(datos.get("confianza", 70))
        if not (0 <= confianza <= 100):
            advertencias.append(f"Confianza fuera de rango: {confianza}")
            datos["confianza"] = max(0, min(100, confianza))  # clamp
    except (ValueError, TypeError):
        datos["confianza"] = 70

    # ── CUFE: alfanumérico si existe ──────────────────────────
    cufe = datos.get("cufe")
    if cufe and cufe not in (None, "null", ""):
        if not re.match(r'^[a-zA-Z0-9\-_]+$', str(cufe)):
            advertencias.append(f"CUFE con caracteres inusuales: '{str(cufe)[:30]}'")

    # ── Verificación de ITBMS ─────────────────────────────────
    advertencias.extend(verificar_itbms(datos))

    # Solo advertencias no críticas → igual es válido
    tiene_error_critico = any(
        "obligatorio" in a or "no es número" in a
        for a in advertencias
    )
    return not tiene_error_critico, advertencias


def validar_resultado(datos):
    """Compatibilidad con el código existente — usa el validador fiscal."""
    if not isinstance(datos, dict):
        return False
    es_valido, advertencias = validar_campos_fiscales(datos)
    for adv in advertencias:
        slog("warning", "Validación fiscal: {}", adv)
    return es_valido


# ─────────────────────────────────────────
def init_db(db_path):
    with sqlite3.connect(db_path) as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS facturas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                archivo TEXT UNIQUE,
                hash TEXT UNIQUE,
                proveedor TEXT,
                ruc TEXT,
                receptor TEXT,
                ruc_receptor TEXT,
                fecha TEXT,
                monto_total REAL,
                subtotal REAL,
                itbms REAL,
                moneda TEXT,
                descripcion TEXT,
                categoria TEXT DEFAULT "Otros",
                confianza INTEGER DEFAULT 70,
                modelo_usado TEXT,
                notas TEXT,
                cufe TEXT,
                tipo_doc TEXT DEFAULT "Factura",
                fuente TEXT DEFAULT "llm",
                estado TEXT DEFAULT "pendiente",
                comentario_estado TEXT,
                fecha_estado TEXT,
                fecha_procesamiento TEXT,
                advertencias_fiscales TEXT
            )
        ''')
        nuevas_columnas = [
            ('categoria',              'TEXT DEFAULT "Otros"'),
            ('receptor',               'TEXT'),
            ('ruc_receptor',           'TEXT'),
            ('subtotal',               'REAL'),
            ('itbms',                  'REAL'),
            ('cufe',                   'TEXT'),
            ('tipo_doc',               'TEXT DEFAULT "Factura"'),
            ('fuente',                 'TEXT DEFAULT "llm"'),
            ('estado',                 'TEXT DEFAULT "pendiente"'),
            ('comentario_estado',      'TEXT'),
            ('fecha_estado',           'TEXT'),
            ('advertencias_fiscales',  'TEXT'),  # ← nueva columna
        ]
        for col, tipo in nuevas_columnas:
            try:
                conn.execute(f'ALTER TABLE facturas ADD COLUMN {col} {tipo}')
                conn.commit()
            except sqlite3.OperationalError:
                pass

# ─────────────────────────────────────────
COLUMNAS_EXCEL = [
    ("archivo",               "Archivo"),
    ("proveedor",             "Proveedor"),
    ("ruc",                   "RUC"),
    ("receptor",              "Receptor"),
    ("ruc_receptor",          "RUC Receptor"),
    ("fecha",                 "Fecha"),
    ("monto_total",           "Monto Total"),
    ("subtotal",              "Subtotal"),
    ("itbms",                 "ITBMS"),
    ("moneda",                "Moneda"),
    ("categoria",             "Categoría"),
    ("tipo_doc",              "Tipo Documento"),
    ("estado",                "Estado"),
    ("comentario_estado",     "Comentario"),
    ("descripcion",           "Descripción"),
    ("cufe",                  "CUFE"),
    ("fuente",                "Fuente"),
    ("confianza",             "Confianza %"),
    ("advertencias_fiscales", "⚠ Advertencias"),   # ← nueva columna en Excel
    ("modelo_usado",          "Modelo"),
    ("fecha_procesamiento",   "Fecha Procesado"),
]

def exportar_excel(db_path, excel_path):
    try:
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            cols_sql = ", ".join(c[0] for c in COLUMNAS_EXCEL)
            rows = conn.execute(
                f"SELECT {cols_sql} FROM facturas ORDER BY fecha_procesamiento DESC"
            ).fetchall()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas"

        header_fill = PatternFill("solid", fgColor="00E5C3")
        header_font = Font(bold=True, color="030F0D")
        for col_idx, (_, label) in enumerate(COLUMNAS_EXCEL, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        warn_fill = PatternFill("solid", fgColor="FEF3C7")  # amarillo suave para advertencias
        for row_idx, row in enumerate(rows, 2):
            for col_idx, (field, _) in enumerate(COLUMNAS_EXCEL, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row[field])
                # Resaltar fila si tiene advertencias fiscales
                if field == "advertencias_fiscales" and row[field]:
                    cell.fill = warn_fill

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        wb.save(excel_path)
        return len(rows)
    except Exception as e:
        slog("error", "Error Excel: {}", str(e))
        return 0


def exportar_dgi_csv(db_path, csv_path, periodo=None):
    """CSV formato DGI Panamá — solo facturas aprobadas."""
    try:
        with sqlite3.connect(db_path) as conn:
            query = """
                SELECT
                    ruc          AS RUC_Proveedor,
                    proveedor    AS Nombre_Proveedor,
                    tipo_doc     AS Tipo_Documento,
                    archivo      AS Numero_Documento,
                    cufe         AS CUFE,
                    fecha        AS Fecha_Emision,
                    COALESCE(subtotal, monto_total - COALESCE(itbms, 0)) AS Subtotal,
                    COALESCE(itbms, 0) AS ITBMS_7pct,
                    monto_total  AS Total_Factura,
                    moneda       AS Moneda,
                    categoria    AS Categoria
                FROM facturas
                WHERE estado = 'aprobada'
            """
            params = []
            if periodo:
                query += " AND fecha LIKE ?"
                params.append(f"{periodo}%")
            query += " ORDER BY fecha ASC"

            rows = conn.execute(query, params).fetchall()
            if not rows:
                return 0

            cols = ["RUC_Proveedor", "Nombre_Proveedor", "Tipo_Documento",
                    "Numero_Documento", "CUFE", "Fecha_Emision",
                    "Subtotal", "ITBMS_7pct", "Total_Factura", "Moneda", "Categoria"]

        import csv
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            f.write("sep=,\n")
            writer = csv.writer(f)
            writer.writerow(cols)
            for row in rows:
                formatted = list(row)
                for i, col in enumerate(cols):
                    if col in ("Subtotal", "ITBMS_7pct", "Total_Factura"):
                        try:
                            formatted[i] = f"{float(row[i] or 0):.2f}"
                        except (TypeError, ValueError):
                            formatted[i] = "0.00"
                writer.writerow(formatted)

        return len(rows)
    except Exception as e:
        slog("error", "Error exportando CSV DGI: {}", str(e))
        return 0

# ─────────────────────────────────────────
def get_file_hash(ruta):
    with open(ruta, "rb") as f:
        return hashlib.md5(f.read()).hexdigest()

def mover_archivo(ruta_origen, carpeta_destino):
    nombre = os.path.basename(ruta_origen)
    for _ in range(6):
        try:
            shutil.move(ruta_origen, os.path.join(carpeta_destino, nombre))
            return True
        except PermissionError:
            time.sleep(1.5)
        except Exception:
            return False
    return False

# ─────────────────────────────────────────
def extraer_datos_xml_etax(ruta_archivo):
    try:
        tree = ET.parse(ruta_archivo)
        root = tree.getroot()
        ns = {
            'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2',
            'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
            'ext': 'urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2',
        }

        def find_text(path, default=''):
            el = root.find(path, ns)
            return el.text.strip() if el is not None and el.text else default

        cufe      = find_text('.//cbc:UUID') or find_text('.//cbc:ID')
        proveedor = (find_text('.//cac:AccountingSupplierParty//cbc:RegistrationName') or
                     find_text('.//cac:AccountingSupplierParty//cbc:Name'))
        ruc       = (find_text('.//cac:AccountingSupplierParty//cbc:CompanyID') or
                     find_text('.//cac:AccountingSupplierParty//cbc:ID'))
        receptor  = (find_text('.//cac:AccountingCustomerParty//cbc:RegistrationName') or
                     find_text('.//cac:AccountingCustomerParty//cbc:Name'))
        ruc_receptor = find_text('.//cac:AccountingCustomerParty//cbc:CompanyID')
        fecha     = find_text('.//cbc:IssueDate')

        monto_total_str = (find_text('.//cac:LegalMonetaryTotal//cbc:PayableAmount') or
                           find_text('.//cbc:TaxInclusiveAmount'))
        subtotal_str    = (find_text('.//cac:LegalMonetaryTotal//cbc:TaxExclusiveAmount') or
                           find_text('.//cac:LegalMonetaryTotal//cbc:LineExtensionAmount'))
        itbms_str       = find_text('.//cac:TaxTotal//cbc:TaxAmount')
        moneda          = find_text('.//cbc:DocumentCurrencyCode') or 'USD'

        tipo_doc_map = {'01': 'Factura', '02': 'Nota de Crédito', '03': 'Nota de Débito',
                        '04': 'Factura de Importación', '08': 'Factura de Exportación'}
        tipo_cod = find_text('.//cbc:InvoiceTypeCode')
        tipo_doc = tipo_doc_map.get(tipo_cod, f'Documento {tipo_cod}')

        descripcion = (find_text('.//cac:InvoiceLine//cbc:Description') or
                       find_text('.//cac:InvoiceLine//cbc:Name') or
                       f'{tipo_doc} electrónica')

        try:
            monto_total = float(monto_total_str)
        except (ValueError, TypeError):
            monto_total = 0.0
        try:
            itbms = float(itbms_str)
        except (ValueError, TypeError):
            itbms = 0.0
        try:
            subtotal = float(subtotal_str)
        except (ValueError, TypeError):
            subtotal = monto_total - itbms

        if not proveedor or monto_total == 0:
            slog("warning", "XML e-Tax sin campos mínimos: {}", ruta_archivo)
            return None

        return {
            'proveedor':    proveedor,
            'ruc':          ruc,
            'receptor':     receptor,
            'ruc_receptor': ruc_receptor,
            'fecha':        fecha,
            'monto_total':  monto_total,
            'subtotal':     subtotal,
            'itbms':        itbms,
            'moneda':       moneda,
            'descripcion':  descripcion,
            'cufe':         cufe,
            'tipo_doc':     tipo_doc,
            'confianza':    100,
            'categoria':    'Otros',
            'notas':        f'e-Tax 2.0 | CUFE: {cufe}',
            'fuente':       'xml_etax',
        }
    except ET.ParseError as e:
        slog("error", "XML inválido {}: {}", ruta_archivo, str(e))
        return None
    except Exception as e:
        slog("error", "Error parseando XML e-Tax {}: {}", ruta_archivo, str(e))
        return None


def extraer_texto(ruta_archivo):
    ext = os.path.splitext(ruta_archivo)[1].lower()
    try:
        if ext == ".txt":
            with open(ruta_archivo, "r", encoding="utf-8", errors="ignore") as f:
                return f.read()

        elif ext == ".pdf":
            # ── PATCH 1: usar pypdf en lugar de PyPDF2 ──────────
            texto = ""
            if _PDF_BACKEND == "pypdf":
                with open(ruta_archivo, "rb") as f:
                    reader = PdfReader(f)
                    for i, page in enumerate(reader.pages):
                        texto += f"\n--- Pagina {i+1} ---\n" + (page.extract_text() or "")
            else:
                # fallback PyPDF2
                with open(ruta_archivo, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    for i, page in enumerate(reader.pages):
                        texto += f"\n--- Pagina {i+1} ---\n" + (page.extract_text() or "")
            return texto

        elif ext in [".xlsx", ".xls"]:
            wb = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
            lineas = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    linea = "\t".join(str(c) if c is not None else "" for c in row)
                    if linea.strip():
                        lineas.append(linea)
            wb.close()
            return "\n".join(lineas)

    except Exception as e:
        slog("error", "Error leyendo {}: {}", ruta_archivo, str(e))
    return None

# ─────────────────────────────────────────
import base64
import mimetypes

MODELOS_VISION = [
    "google/gemini-2.5-flash-preview:free",
    "google/gemini-2.0-flash-lite-001",
    "meta-llama/llama-3.2-11b-vision-instruct:free",
    "google/gemini-flash-1.5-8b",
]

def extraer_datos_imagen(ruta_archivo):
    """
    Envía la imagen directamente a un modelo vision de OpenRouter.
    Retorna (datos_dict, nombre_modelo) o (None, None) si falla.
    """
    if not API_KEY:
        slog("error", "OPENROUTER_API_KEY no configurada para vision")
        return None, None

    ext = os.path.splitext(ruta_archivo)[1].lower()
    mime_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png", ".webp": "image/webp"}
    mime_type = mime_map.get(ext, "image/jpeg")

    try:
        with open(ruta_archivo, "rb") as f:
            img_b64 = base64.b64encode(f.read()).decode("utf-8")
    except Exception as e:
        slog("error", "No se pudo leer imagen {}: {}", ruta_archivo, str(e))
        return None, None

    prompt_texto = (
        "Eres un asistente contable experto en facturas de Panama y Latinoamerica.\n"
        "Analiza la imagen de esta factura y extrae los datos visibles.\n\n"
        "INSTRUCCIONES:\n"
        "- Responde UNICAMENTE con un JSON valido, sin texto antes ni despues\n"
        "- Si no encuentras un campo, usa null\n"
        "- monto_total debe ser numero, no texto\n"
        "- itbms es el impuesto del 7% en Panama\n"
        "- fecha en formato DD/MM/AAAA\n"
        "- moneda: USD, PAB, o la que veas\n"
        "- categoria debe ser UNA de: Servicios, Materiales y Suministros, "
        "Transporte y Logistica, Tecnologia y Software, Nomina y RRHH, "
        "Alquiler e Inmuebles, Publicidad y Marketing, Impuestos y Tasas, Alimentacion, Otros\n\n"
        "FORMATO EXACTO:\n"
        '{"proveedor": "nombre", "ruc": "ruc o null", '
        '"fecha": "DD/MM/AAAA o null", "monto_total": 0.00, '
        '"subtotal": 0.00, "itbms": 0.00, '
        '"moneda": "USD", "categoria": "Servicios", '
        '"descripcion": "breve", "confianza": 85, "notas": null}'
    )

    def _llamar_vision(model):
        nombre = model.split('/')[1].split(':')[0] if '/' in model else model
        try:
            data = {
                "model": model,
                "max_tokens": 500,
                "temperature": 0.0,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{img_b64}"}},
                        {"type": "text",      "text": prompt_texto}
                    ]
                }]
            }
            req = urllib.request.Request(
                url="https://openrouter.ai/api/v1/chat/completions",
                data=json.dumps(data).encode("utf-8"),
                headers={
                    "Content-Type": "application/json",
                    "Authorization": f"Bearer {API_KEY}",
                    "HTTP-Referer": "https://automatiapa.pythonanywhere.com",
                    "X-Title": "AutomatIA Facturas Vision"
                },
                method="POST"
            )
            with urllib.request.urlopen(req, timeout=45) as resp:
                result = json.loads(resp.read().decode())
                if "error" in result:
                    slog("warning", "Vision {}: API error: {}", nombre, str(result["error"]))
                    return None, nombre
                if "choices" not in result or not result["choices"]:
                    slog("warning", "Vision {}: sin choices en respuesta", nombre)
                    return None, nombre
                resp_content = result["choices"][0]["message"]["content"].strip()
                datos = parsear_json_respuesta(resp_content)
                if datos and validar_resultado(datos):
                    datos["fuente"] = "vision"
                    return datos, nombre
                slog("warning", "Vision {}: JSON invalido o campos faltantes — resp: {}", nombre, resp_content[:200])
                return None, nombre
        except urllib.error.HTTPError as e:
            body = e.read().decode("utf-8", errors="ignore")[:300]
            slog("warning", "Vision {} HTTP {}: {}", nombre, str(e.code), body)
            return None, nombre
        except urllib.error.URLError as e:
            slog("warning", "Vision {} URLError (dominio bloqueado o sin red): {}", nombre, str(e.reason))
            return None, nombre
        except Exception as e:
            slog("warning", "Vision {}: {}", nombre, str(e))
            return None, nombre

    # Llamadas secuenciales — mas seguro en PythonAnywhere (evita limites de threads)
    for model in MODELOS_VISION:
        datos, nombre = _llamar_vision(model)
        if datos:
            slog("info", "Vision OK: {} con {}", os.path.basename(ruta_archivo), nombre)
            return datos, nombre

    slog("error", "Todos los modelos vision fallaron: {}", ruta_archivo)
    return None, None

# ─────────────────────────────────────────
def parsear_json_respuesta(content):
    if "<think>" in content:
        fin_think = content.find("</think>")
        if fin_think != -1:
            content = content[fin_think + len("</think>"):].strip()
    if "```" in content:
        partes = content.split("```")
        for parte in partes:
            parte = parte.replace("json", "").strip()
            if "{" in parte:
                content = parte
                break
    inicio = content.find("{")
    fin    = content.rfind("}") + 1
    if inicio == -1 or fin <= inicio:
        return None
    return json.loads(content[inicio:fin])

# ─────────────────────────────────────────
from concurrent.futures import ThreadPoolExecutor, as_completed

def _llamar_modelo(model, prompt, api_key):
    """Llama a un modelo específico. Retorna (datos, nombre_modelo) o (None, None)."""
    nombre = model.split('/')[1].split(':')[0]
    try:
        data = {
            "model": model,
            "messages": [{"role": "user", "content": prompt}],
            "temperature": 0.0,
            "max_tokens": 1500
        }
        req = urllib.request.Request(
            url="https://openrouter.ai/api/v1/chat/completions",
            data=json.dumps(data).encode("utf-8"),
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}",
                "HTTP-Referer": "https://automatiapa.pythonanywhere.com",
                "X-Title": "AutomatIA Facturas"
            },
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read().decode())
            if "choices" not in result or not result["choices"]:
                return None, nombre
            content = result["choices"][0]["message"]["content"].strip()
            if not content:
                return None, nombre
            datos = parsear_json_respuesta(content)
            if datos and validar_resultado(datos):
                return datos, nombre
            return None, nombre
    except Exception as e:
        slog("warning", "{}: error {}", nombre, str(e))
        return None, nombre


def llamar_ia(texto):
    # Preprocesar texto: quedarse solo con las primeras líneas útiles
    lineas = [l.strip() for l in texto.splitlines() if l.strip()]
    texto_limpio = "\n".join(lineas)[:12000].replace('"', "'").replace('\\', '/')

    prompt = (
        "Eres un asistente contable experto en facturas de Panama y Latinoamerica.\n"
        "Analiza el siguiente texto de una factura y extrae los datos.\n\n"
        "INSTRUCCIONES IMPORTANTES:\n"
        "- Responde UNICAMENTE con un JSON valido, sin texto antes ni despues\n"
        "- No uses bloques de codigo ni comillas triples\n"
        "- Si no encuentras un campo, usa null\n"
        "- monto_total debe ser un numero, no texto\n"
        "- subtotal es el monto antes de impuestos (sin ITBMS)\n"
        "- itbms es el impuesto de transferencia (7% en Panama)\n"
        "- La relacion esperada: subtotal + itbms = monto_total\n"
        "- Si la factura no desglosa impuesto, usa null en subtotal e itbms\n"
        "- fecha debe estar en formato DD/MM/AAAA\n"
        "- moneda: usa USD, PAB, o la que corresponda\n"
        "- categoria debe ser UNA de estas: "
        "Servicios, Materiales y Suministros, Transporte y Logistica, "
        "Tecnologia y Software, Nomina y RRHH, Alquiler e Inmuebles, "
        "Publicidad y Marketing, Impuestos y Tasas, Alimentacion, Otros\n\n"
        "FORMATO EXACTO:\n"
        '{"proveedor": "nombre", "ruc": "ruc o null", '
        '"fecha": "DD/MM/AAAA o null", "monto_total": 0.00, '
        '"subtotal": 0.00, "itbms": 0.00, '
        '"moneda": "USD", "categoria": "Servicios", '
        '"descripcion": "breve", "confianza": 85, "notas": null}\n\n'
        "FACTURA:\n" + texto_limpio
    )

    # Ronda 1: los 3 modelos más rápidos en paralelo
    ronda_1 = MODELOS[:3]
    # Ronda 2: el resto en paralelo si la primera falla
    ronda_2 = MODELOS[3:]

    for ronda, modelos in [("1", ronda_1), ("2", ronda_2)]:
        print(f"   Ronda {ronda}: probando {len(modelos)} modelos en paralelo...")
        with ThreadPoolExecutor(max_workers=len(modelos)) as executor:
            futuros = {
                executor.submit(_llamar_modelo, m, prompt, API_KEY): m
                for m in modelos
            }
            for futuro in as_completed(futuros):
                datos, nombre = futuro.result()
                if datos:
                    print(f"   ✓ Respondió: {nombre} (confianza: {datos.get('confianza')}%)")
                    return datos, nombre

    slog("error", "Todos los modelos fallaron para esta factura")
    return None, None

# ─────────────────────────────────────────
def guardar_factura(db_path, datos, archivo, modelo):
    categoria = datos.get('categoria', 'Otros')
    if categoria not in CATEGORIAS_VALIDAS:
        categoria = 'Otros'

    # ── PATCH 3: guardar advertencias fiscales en la DB ──────
    _, advertencias = validar_campos_fiscales(datos)

    # ── Detección de duplicados por contenido ────────────────
    # Igual proveedor (ignora mayúsculas) + mismo monto + misma fecha
    # → avisa sin bloquear, para que el usuario lo revise
    try:
        proveedor_nuevo = (datos.get('proveedor') or '').strip().lower()
        monto_nuevo     = round(float(datos.get('monto_total') or 0), 2)
        fecha_nueva     = (datos.get('fecha') or '').strip()

        if proveedor_nuevo and monto_nuevo and fecha_nueva:
            with sqlite3.connect(db_path) as conn:
                dup = conn.execute("""
                    SELECT archivo FROM facturas
                    WHERE LOWER(TRIM(proveedor)) = ?
                      AND ROUND(monto_total, 2) = ?
                      AND fecha = ?
                    LIMIT 1
                """, (proveedor_nuevo, monto_nuevo, fecha_nueva)).fetchone()
            if dup:
                advertencias.append(
                    f"⚠ POSIBLE DUPLICADO: misma combinación proveedor/monto/fecha "
                    f"ya existe en '{dup[0]}'"
                )
                slog("warning", "Posible duplicado por contenido: {} ≈ {}", archivo, dup[0])
    except Exception as e:
        slog("warning", "Error en detección de duplicados: {}", str(e))

    advertencias_str = " | ".join(advertencias) if advertencias else None

    try:
        with sqlite3.connect(db_path) as conn:
            conn.execute('''
                INSERT INTO facturas
                (archivo, hash, proveedor, ruc, receptor, ruc_receptor, fecha,
                 monto_total, subtotal, itbms, moneda, descripcion, categoria,
                 confianza, modelo_usado, notas, cufe, tipo_doc, fuente,
                 estado, fecha_procesamiento, advertencias_fiscales)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'pendiente', ?, ?)
            ''', (
                archivo,
                datos.get('hash'),
                datos.get('proveedor'),
                datos.get('ruc'),
                datos.get('receptor'),
                datos.get('ruc_receptor'),
                datos.get('fecha'),
                float(datos.get('monto_total', 0)),
                float(datos.get('subtotal', 0)) if datos.get('subtotal') else None,
                float(datos.get('itbms', 0))    if datos.get('itbms')    else None,
                datos.get('moneda', 'USD'),
                datos.get('descripcion'),
                categoria,
                int(datos.get('confianza', 70)),
                modelo,
                datos.get('notas'),
                datos.get('cufe'),
                datos.get('tipo_doc', 'Factura'),
                datos.get('fuente', 'llm'),
                datetime.now().isoformat(),
                advertencias_str,
            ))
        return True
    except sqlite3.IntegrityError:
        slog("warning", "Factura duplicada: {}", archivo)
        return False
    except Exception as e:
        slog("error", "Error guardando factura {}: {}", archivo, str(e))
        return False

# ══════════════════════════════════════════════════════════════
# FORMULARIO 103 — DECLARACIÓN JURADA DE COMPRAS Y GASTOS
#
# Genera un Excel con el formato que exige la DGI Panamá para el
# Formulario 103 (Retención de ITBMS en compras a no residentes
# y declaración de gastos deducibles).
#
# Estructura basada en el formulario público DGI F-103 v2020:
#   Sección A — Identificación del declarante
#   Sección B — Detalle de compras por proveedor (una fila por RUC)
#   Sección C — Totales
# ══════════════════════════════════════════════════════════════
def exportar_formulario103(db_path: str, output_path: str,
                           nombre_empresa: str, ruc_empresa: str = "",
                           periodo: str = "") -> dict:
    """
    Genera el Formulario 103 precompletado como archivo Excel.

    Args:
        db_path:        Ruta a la base de datos SQLite del cliente.
        output_path:    Ruta de salida del archivo .xlsx.
        nombre_empresa: Nombre o razón social del declarante.
        ruc_empresa:    RUC del declarante (opcional).
        periodo:        'YYYY-MM' para filtrar por mes, o '' para todo.

    Returns:
        dict con claves: total_proveedores, total_compras,
                         total_itbms, total_retencion, ruta
    """
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                 Border, Side, numbers)
    from openpyxl.utils import get_column_letter

    VERDE_DGI  = "1A5276"   # azul oscuro DGI (encabezado)
    GRIS_SUAVE = "D5D8DC"
    AMARILLO   = "FCF3CF"
    BLANCO     = "FFFFFF"

    thin = Side(style="thin", color="AAAAAA")
    borde = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_cell(ws, row, col, value, bg=VERDE_DGI, fg="FFFFFF", bold=True, size=10, wrap=False):
        c = ws.cell(row=row, column=col, value=value)
        c.fill   = PatternFill("solid", fgColor=bg)
        c.font   = Font(bold=bold, color=fg, size=size)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border = borde
        return c

    def data_cell(ws, row, col, value, fmt=None, bg=BLANCO, bold=False, align="left"):
        c = ws.cell(row=row, column=col, value=value)
        c.fill   = PatternFill("solid", fgColor=bg)
        c.font   = Font(bold=bold, size=10)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = borde
        if fmt:
            c.number_format = fmt
        return c

    # ── Leer facturas aprobadas de la DB ──────────────────────
    try:
        with sqlite3.connect(db_path) as conn:
            conn.row_factory = sqlite3.Row
            q = """
                SELECT
                    ruc, proveedor,
                    SUM(monto_total)                              AS total_compras,
                    SUM(COALESCE(subtotal,
                        monto_total - COALESCE(itbms, 0)))        AS total_subtotal,
                    SUM(COALESCE(itbms, 0))                       AS total_itbms,
                    COUNT(*)                                      AS num_facturas,
                    MAX(fecha)                                    AS ultima_fecha,
                    categoria
                FROM facturas
                WHERE estado = 'aprobada'
            """
            params = []
            if periodo:
                q += " AND fecha LIKE ?"
                params.append(f"{periodo}%")
            q += " GROUP BY COALESCE(ruc, proveedor) ORDER BY total_compras DESC"
            rows = conn.execute(q, params).fetchall()
    except Exception as e:
        slog("error", "F103 error leyendo DB: {}", str(e))
        return {"error": str(e)}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Formulario 103"

    # ── ENCABEZADO INSTITUCIONAL ──────────────────────────────
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "REPÚBLICA DE PANAMÁ — MINISTERIO DE ECONOMÍA Y FINANZAS — DIRECCIÓN GENERAL DE INGRESOS"
    c.font  = Font(bold=True, size=11, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor=VERDE_DGI)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value = "FORMULARIO 103 — DECLARACIÓN JURADA DE COMPRAS Y GASTOS (ITBMS)"
    c.font  = Font(bold=True, size=12, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor=VERDE_DGI)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22

    # ── SECCIÓN A — IDENTIFICACIÓN DEL DECLARANTE ─────────────
    ROW_A = 4
    ws.merge_cells(f"A{ROW_A}:J{ROW_A}")
    c = ws.cell(row=ROW_A, column=1, value="SECCIÓN A — IDENTIFICACIÓN DEL DECLARANTE")
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2E86C1")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[ROW_A].height = 16

    # Fila datos declarante
    etiquetas_a = [
        ("Nombre / Razón Social:", nombre_empresa),
        ("RUC:",                   ruc_empresa or "—"),
        ("Período:",               periodo or "Todos"),
        ("Fecha generación:",      datetime.now().strftime("%d/%m/%Y")),
    ]
    col_a = 1
    for etiq, val in etiquetas_a:
        ws.cell(row=ROW_A+1, column=col_a,   value=etiq).font = Font(bold=True, size=9)
        ws.cell(row=ROW_A+1, column=col_a+1, value=val).font  = Font(size=9)
        col_a += 3
    ws.row_dimensions[ROW_A+1].height = 14

    # ── SECCIÓN B — DETALLE POR PROVEEDOR ─────────────────────
    ROW_B = ROW_A + 3
    ws.merge_cells(f"A{ROW_B}:J{ROW_B}")
    c = ws.cell(row=ROW_B, column=1, value="SECCIÓN B — DETALLE DE COMPRAS POR PROVEEDOR / SUPLIDOR")
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2E86C1")
    c.alignment = Alignment(horizontal="left", vertical="center")

    COLS_B = [
        ("N°",                  5),
        ("RUC del Proveedor",  18),
        ("Nombre / Razón Social", 32),
        ("Categoría",          18),
        ("N° Facturas",        12),
        ("Última Fecha",       14),
        ("Base Imponible\n(Subtotal)", 16),
        ("ITBMS\nDeclarado",   14),
        ("Retención\n(1% ITBMS)", 14),
        ("Total\nCompras",     14),
    ]
    ROW_HDR = ROW_B + 1
    for ci, (label, _) in enumerate(COLS_B, 1):
        hdr_cell(ws, ROW_HDR, ci, label, wrap=True)
    ws.row_dimensions[ROW_HDR].height = 30

    # Anchos de columna
    for ci, (_, w) in enumerate(COLS_B, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Filas de datos
    total_compras   = 0.0
    total_itbms_sum = 0.0
    total_subtotal  = 0.0
    total_retencion = 0.0

    for idx, row in enumerate(rows, 1):
        r = ROW_HDR + idx
        bg = BLANCO if idx % 2 == 0 else GRIS_SUAVE

        tc  = float(row["total_compras"]  or 0)
        ti  = float(row["total_itbms"]    or 0)
        ts  = float(row["total_subtotal"] or 0)
        ret = round(ti * 0.01, 2)   # retención del 1% sobre ITBMS pagado

        total_compras   += tc
        total_itbms_sum += ti
        total_subtotal  += ts
        total_retencion += ret

        data_cell(ws, r, 1,  idx,                            align="center", bg=bg)
        data_cell(ws, r, 2,  row["ruc"] or "SIN RUC",       bg=bg)
        data_cell(ws, r, 3,  row["proveedor"] or "",         bg=bg)
        data_cell(ws, r, 4,  row["categoria"] or "Otros",    bg=bg)
        data_cell(ws, r, 5,  row["num_facturas"],            align="center", bg=bg)
        data_cell(ws, r, 6,  row["ultima_fecha"] or "",      align="center", bg=bg)
        data_cell(ws, r, 7,  ts,  fmt='#,##0.00 "B/."',     align="right", bg=bg)
        data_cell(ws, r, 8,  ti,  fmt='#,##0.00 "B/."',     align="right", bg=bg)
        data_cell(ws, r, 9,  ret, fmt='#,##0.00 "B/."',     align="right", bg=bg)
        data_cell(ws, r, 10, tc,  fmt='#,##0.00 "B/."',     align="right", bg=bg)

    # ── SECCIÓN C — TOTALES ───────────────────────────────────
    ROW_TOT = ROW_HDR + len(rows) + 2
    ws.merge_cells(f"A{ROW_TOT}:J{ROW_TOT}")
    c = ws.cell(row=ROW_TOT, column=1, value="SECCIÓN C — TOTALES")
    c.font = Font(bold=True, size=10, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor="2E86C1")
    c.alignment = Alignment(horizontal="left", vertical="center")

    tot_labels = [
        ("Total Proveedores Declarados:", len(rows),            None),
        ("Total Base Imponible:",          total_subtotal,  '#,##0.00 "B/."'),
        ("Total ITBMS Declarado:",         total_itbms_sum, '#,##0.00 "B/."'),
        ("Total Retención (1%):",          total_retencion, '#,##0.00 "B/."'),
        ("TOTAL GENERAL COMPRAS:",         total_compras,   '#,##0.00 "B/."'),
    ]
    for ti_idx, (label, val, fmt) in enumerate(tot_labels):
        r = ROW_TOT + 1 + ti_idx
        ws.merge_cells(f"A{r}:F{r}")
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(bold=True, size=10)
        lc.fill = PatternFill("solid", fgColor=AMARILLO)
        lc.alignment = Alignment(horizontal="right", vertical="center")
        lc.border = borde

        ws.merge_cells(f"G{r}:J{r}")
        vc = ws.cell(row=r, column=7, value=val)
        vc.font = Font(bold=True, size=10)
        vc.fill = PatternFill("solid", fgColor=AMARILLO)
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.border = borde
        if fmt:
            vc.number_format = fmt

    # ── NOTA LEGAL ────────────────────────────────────────────
    ROW_NOTA = ROW_TOT + len(tot_labels) + 2
    ws.merge_cells(f"A{ROW_NOTA}:J{ROW_NOTA}")
    nota = (
        "NOTA: Este formulario es generado automáticamente por AutomatIA a partir de las facturas "
        "aprobadas en el sistema. Verifique los datos antes de presentar ante la DGI. "
        "La retención del 1% aplica según el Art. 1057-V del Código Fiscal."
    )
    nc = ws.cell(row=ROW_NOTA, column=1, value=nota)
    nc.font = Font(italic=True, size=8, color="555555")
    nc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[ROW_NOTA].height = 30

    try:
        wb.save(output_path)
    except Exception as e:
        slog("error", "F103 error guardando Excel: {}", str(e))
        return {"error": str(e)}

    return {
        "total_proveedores": len(rows),
        "total_compras":     round(total_compras, 2),
        "total_itbms":       round(total_itbms_sum, 2),
        "total_retencion":   round(total_retencion, 2),
        "ruta":              output_path,
    }


# ─────────────────────────────────────────
def procesar_cliente(nombre_cliente, procesadas):
    rutas = get_rutas_cliente(nombre_cliente)
    crear_carpetas_cliente(rutas)
    init_db(rutas["db"])

    archivos = [
        f for f in os.listdir(rutas["facturas"])
        if f.lower().endswith(('.pdf', '.txt', '.xlsx', '.xls', '.xml', '.jpg', '.jpeg', '.png', '.webp'))
    ]

    clave = lambda f: f"{nombre_cliente}/{f}"

    for archivo in archivos:
        if clave(archivo) in procesadas:
            continue

        ruta = os.path.join(rutas["facturas"], archivo)
        ext  = os.path.splitext(archivo)[1].lower()
        print(f"\n  [{_sanitize_log(nombre_cliente)}] Procesando: {_sanitize_log(archivo)}")
        slog("info", "Iniciando: {}/{}", nombre_cliente, archivo)

        file_hash = get_file_hash(ruta)

        with sqlite3.connect(rutas["db"]) as conn:
            existe = conn.execute(
                "SELECT 1 FROM facturas WHERE hash=?", (file_hash,)
            ).fetchone()

        if existe:
            print("   Duplicada, ignorando")
            mover_archivo(ruta, rutas["procesados"])
            procesadas.add(clave(archivo))
            continue

        resultado = None
        modelo    = None

        if ext == ".xml":
            resultado = extraer_datos_xml_etax(ruta)
            if resultado:
                modelo = "e-Tax-XML"
                resultado['hash'] = file_hash
            else:
                print("   XML inválido o sin campos mínimos")
                slog("error", "XML e-Tax inválido: {}", archivo)
                mover_archivo(ruta, rutas["error"])
                procesadas.add(clave(archivo))
                continue
        elif ext in (".jpg", ".jpeg", ".png", ".webp"):
            print(f"   Imagen detectada — usando modelo vision...")
            resultado, modelo = extraer_datos_imagen(ruta)
            if resultado:
                resultado['hash'] = file_hash
            else:
                print("   Modelos vision fallaron")
                slog("error", "Vision fallida: {}", archivo)
                mover_archivo(ruta, rutas["error"])
                procesadas.add(clave(archivo))
                continue
        else:
            texto = extraer_texto(ruta)
            if not texto or len(texto.strip()) < 20:
                print("   No se pudo extraer texto util")
                slog("error", "Texto insuficiente: {}", archivo)
                mover_archivo(ruta, rutas["error"])
                procesadas.add(clave(archivo))
                continue

            resultado, modelo = llamar_ia(texto)
            if resultado:
                resultado['hash'] = file_hash

        if resultado:
            if guardar_factura(rutas["db"], resultado, archivo, modelo):
                fuente_tag = "🧾 e-Tax XML" if ext == ".xml" else f"🤖 {modelo}"
                print(f"   OK: {resultado.get('proveedor')} | {resultado.get('monto_total')} {resultado.get('moneda')} | {fuente_tag}")
                slog("info", "OK: {} | {} | {}", archivo, str(resultado.get('proveedor')), str(resultado.get('monto_total')))
                enviar_telegram(
                    f"{'🧾' if ext == '.xml' else '📄'} Factura procesada\n"
                    f"Cliente: {nombre_cliente}\n"
                    f"Archivo: {archivo}\n"
                    f"Fuente: {fuente_tag}\n"
                    f"Proveedor: {resultado.get('proveedor')}\n"
                    f"Categoría: {resultado.get('categoria', 'Otros')}\n"
                    f"Monto: {resultado.get('monto_total')} {resultado.get('moneda')}\n"
                    f"Confianza: {resultado.get('confianza')}%"
                )
                mover_archivo(ruta, rutas["procesados"])
            else:
                print("   Ya estaba guardada (duplicado por hash)")
                mover_archivo(ruta, rutas["procesados"])
        else:
            print("   Fallo el procesamiento con todos los modelos")
            slog("error", "FALLO TOTAL: {}/{}", nombre_cliente, archivo)
            enviar_telegram(
                f"❌ Error procesando factura\n"
                f"Cliente: {nombre_cliente}\n"
                f"Archivo: {archivo}\n"
                f"Todos los modelos fallaron"
            )
            mover_archivo(ruta, rutas["error"])

        procesadas.add(clave(archivo))

    exportar_excel(rutas["db"], rutas["excel"])

# ─────────────────────────────────────────
def main():
    os.makedirs(CARPETA_CLIENTES, exist_ok=True)
    print("=" * 60)
    print("  SISTEMA MULTI-CLIENTE DE FACTURAS")
    print(f"  PDF backend: {_PDF_BACKEND}")
    print("=" * 60)
    enviar_telegram("Sistema multi-cliente iniciado")
    procesadas = set()
    while True:
        clientes = listar_clientes()
        if not clientes:
            print(" Sin clientes aun. Crea una carpeta en ./clientes/", end="\r")
        else:
            for cliente in clientes:
                procesar_cliente(cliente, procesadas)
        time.sleep(15)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nSistema detenido.")
    except Exception as e:
        slog("error", "Error critico: {}", str(e))
        print(f"Error critico: {e}")
